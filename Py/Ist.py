#  Imports
import pandas as pd
from openpyxl import load_workbook

#  Global variable
gcount = 1


#  Class
class Read:
    #  Validation Class to validate Name, PS Number, Email
    def validation(self, ID, Name, E_mail):
                  count = 0  # Local Count
        wb = load_workbook('D:\Py\Book1.xlsx')
    sheets = wb.sheetnames
        for i in range(len(wb.worksheets)):
            sheet = wb[sheets[i]]
            for j in range(2, sheet.max_row + 1):
                if sheet.cell(row=j, column=1).value == ID and sheet.cell(row=j,column=2).value == Name and sheet.cell(row=j, column=3).value == E_mail:
                    count += 1
                    break
        if count == 0:
            print("\nData Provided NOT FOUND in DataBase\n")
            globals()['gcount'] = 0
            wb.close()
        else:
            print("\nData Present in Database\n")
            globals()['gcount'] = 1
            wb.close()

    #  Our main function which will Read from Excel and Write in Master Sheet and Summary

    def ReadWrite(self, ps):

        #   WorkBook Load (Sheets Loading in List df)

        sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
        df = []

        for i in range(0, 5):
            df.append(pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=sheets[i]))

        # Validation (PS Number) from all sheets and appending it in a dataframe

        df1 = pd.DataFrame()
        for i in range(0, 5):
            up_d = df[i].loc[(df[i]['Ps No'] == ps)]
            df1 = df1.append(up_d)

        # a1 = pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=0)
        # a2 = pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=1)
        # a3 = pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=2)
        # a4 = pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=3)
        # a5 = pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name=4)
        #
        # # Merging all data  from sheets into a single row
        # a6 = a1.merge(a2, on="Parent ID", how="left")
        # a7 = a6.merge(a3, on="Patient ID", how="left")
        # a8 = a7.merge(a4, on="Patient_ID", how="left")
        # d = a8.merge(a5, on="Patient ID", how="left")

        d = {"ID","Name ","E_mail ","Company ","Blood Group ","Age ","Gender","Permanent  Address ","Father Name ","Mother name ","Brother Name "," Sister name","Married or Not","College Name ","School Name ","Current Address ","Weight ","Height ","Any Kind of Disability"," Mobile no","Private E-mail ","Adhaar Card No ","Pan Card No ","Favourite Sports ","Bank Account Name","Facebook Account ","Linkedin Profile ","Instagram Account ","Habit ","Nationality ","Religion","Git Hub Account  ","Break Hrs ","High Qualification ","Profile ","Sailary ","In Time"," Out Time"}

        df1 = df1.groupby('ID', as_index=False).aggregate(d).reindex(columns=df1.columns)

        # Using Openpyxl to create and load dataframe to workbook

        book = load_workbook(r"D:\Py\Book1.xlsx")
        writer = pd.ExcelWriter(r"D:\Py\Book1.xlsx", engine='openpyxl')
        writer.book = book
        writer.sheets = dict(
            (ws.title, ws) for ws in book.worksheets)  # to append data on MasterSheet and not create another Sheet

        # condition to create New sheet or Append if MasterSheet already created.

        sheets = book.sheetnames
        if 'MasterSheet' in sheets:
            print("Master Sheet present")
            sheet = book['MasterSheet']
            df1.to_excel(writer, sheet_name='MasterSheet', index=False, header=False, startrow=sheet.max_row)

        else:
            df1.to_excel(writer, sheet_name='MasterSheet', index=False)  # This will create new MasterSheet

        #  Save the Excel File and Print the Updated sheet
        sheet = book['MasterSheet']
        data = {'Number of Trainers': [sheet.max_row - 1],
                'Individual Data': [sheet.max_column],
                'Total Data': [(sheet.max_row - 1) * sheet.max_column],
                }

        df2 = pd.DataFrame(data, columns=['Number of Trainers', 'Individual Data', 'Total Data'])
        df2.to_excel(writer, sheet_name='Summary', index=False)  # This will create new Summary

        book.save("D:\Py\Book1.xlsx")
        print(pd.read_excel(r'D:\Py\Book1.xlsx', sheet_name='MasterSheet'))

        # Excel File Close

        book.close()


d1 = Read()
no_of_inputs = int(input("Select the number of inputs: "))
for i in range(no_of_inputs):
    Name = input("Enter the name for Data" + str(i) + " : ")
    ID = int(input("Enter the PS No for Data" + str(i) + " : "))
    E_mail  = input("Enter email id for Data" + str(i) + " : ")
    d1.validation(ID,Name, E_mail)
    if gcount == 0:
        continue
    d1.ReadWrite(ID)
